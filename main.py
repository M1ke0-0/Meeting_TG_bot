import asyncio
import logging
import sys
import re
import sqlite3
from openpyxl import load_workbook
from datetime import datetime

from aiogram import Bot, Dispatcher, types
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram import F
from aiogram import BaseMiddleware
from aiogram.types import TelegramObject

from dotenv import load_dotenv
import os
import uuid

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")

if not BOT_TOKEN:
    print("Ошибка: переменная BOT_TOKEN не найдена в файле .env")
    print("Создайте файл .env и добавьте строку:")
    print("BOT_TOKEN=ваш_токен_от_BotFather")
    sys.exit(1)

DB_PATH = "bot.db"

class Registration(StatesGroup):
    name = State()
    surname = State()
    gender = State()
    age = State()
    region = State()
    interests = State()
    photo = State()
    location = State()
    

class AdminLoad(StatesGroup):
    waiting_excel = State()


class CreateEvent(StatesGroup):
    name = State()
    date = State()
    time = State()
    interests = State()
    address = State()
    description = State()
    photo = State()
    invite_friends = State()
    confirm = State()


def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            number          TEXT PRIMARY KEY,
            role            TEXT NOT NULL CHECK(role IN ('admin', 'user')),
            registered      INTEGER DEFAULT 0,
            tg_id           INTEGER,
            name            TEXT,
            surname         TEXT,
            gender          TEXT,
            age             INTEGER,
            region          TEXT,
            interests       TEXT,
            photo_file_id   TEXT,
            document_file_id TEXT, 
            location_lat    REAL,
            location_lon    REAL,
            created_at      DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            organizer_phone TEXT NOT NULL,
            name TEXT NOT NULL,
            date TEXT NOT NULL,          -- ДД.ММ.ГГГГ
            time TEXT NOT NULL,          -- ЧЧ:ММ
            interests TEXT,              -- через запятую
            address TEXT,
            description TEXT,
            photo_file_id TEXT,
            document_file_id TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (organizer_phone) REFERENCES users(number)
        )
    ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS event_participants (
            event_id INTEGER,
            participant_phone TEXT,
            joined_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (event_id, participant_phone),
            FOREIGN KEY (event_id) REFERENCES events(id),
            FOREIGN KEY (participant_phone) REFERENCES users(number)
        )
    ''')
        
    conn.commit()
    conn.close()
    logging.info("Таблица users готова")





def init_admin_tables():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS interests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS regions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
    """)

    conn.commit()
    conn.close()


def check_user_status(phone: str) -> dict:
    """Возвращает статус пользователя или None"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT role, registered, tg_id, name
            FROM users WHERE number = ?
        """, (phone,))
        row = c.fetchone()
    
    if row:
        return {
            "exists": True,
            "role": row[0],
            "registered": bool(row[1]),
            "tg_id": row[2],
            "name": row[3]
        }
    return {"exists": False}

def register_phone(phone: str, tg_id: int):
    """Добавляем только номер + tg_id, роль по умолчанию 'user'"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        try:
            c.execute("""
                INSERT INTO users (number, role, tg_id, registered)
                VALUES (?, 'user', ?, 0)
            """, (phone, tg_id))
            conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False


def update_user_profile(phone: str, data: dict):
    """Обновляем профиль пользователя в БД"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            UPDATE users SET
                name = ?, surname = ?, gender = ?, age = ?,
                region = ?, interests = ?, photo_file_id = ?,
                document_file_id = ?, location_lat = ?, location_lon = ?,
                registered = 1
            WHERE number = ?
        """, (
            data.get('name'), data.get('surname'), data.get('gender'),
            data.get('age'), data.get('region'),
            ','.join(data.get('interests', [])) if data.get('interests') else None,
            data.get('photo_file_id'), data.get('document_file_id'),
            data.get('location_lat'), data.get('location_lon'), phone
        ))
        conn.commit()


def replace_interests(interests: list[str]):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("DELETE FROM interests")
        c.executemany(
            "INSERT OR IGNORE INTO interests (name) VALUES (?)",
            [(i.strip(),) for i in interests if i.strip()]
        )
        conn.commit()



def replace_regions(regions: list[str]):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("DELETE FROM regions")
        c.executemany(
            "INSERT OR IGNORE INTO regions (name) VALUES (?)",
            [(r.strip(),) for r in regions if r.strip()]
        )
        conn.commit()

def get_all_regions() -> list[str]:
    """Возвращает актуальный список регионов из БД"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT name FROM regions ORDER BY name")
        regions = [row[0] for row in c.fetchall()]
    return regions if regions else ["Регионы пока не добавлены"]


def get_all_interests() -> list[str]:
    """Возвращает актуальный список интересов из БД"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("SELECT name FROM interests ORDER BY name")
        interests = [row[0] for row in c.fetchall()]
    return interests if interests else ["Интересы пока не добавлены"]


def get_user_by_tg_id(tg_id: int) -> dict | None:
    """Получает все данные пользователя по telegram ID"""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT number, role, registered, name, surname, gender, age, 
                   region, interests, photo_file_id, document_file_id, 
                   location_lat, location_lon
            FROM users WHERE tg_id = ?
        """, (tg_id,))
        row = c.fetchone()
    
    if row:
        columns = [
            "number", "role", "registered", "name", "surname", "gender",
            "age", "region", "interests", "photo_file_id", "document_file_id",
            "location_lat", "location_lon"
        ]
        user = dict(zip(columns, row))
        user["registered"] = bool(user["registered"])
        return user
    return None



def get_event_card_keyboard_optimized(event_id: int, user_phone: str, 
                                     organizer_phone: str, is_participant: bool):
    """Клавиатура для карточки мероприятия (БЕЗ запроса к БД)"""
    if user_phone == organizer_phone:
        return None
    
    if is_participant:
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="❌ Отказаться от участия", 
                                callback_data=f"leave_event_{event_id}")]
        ])
    else:
        return InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Участвовать", 
                                callback_data=f"join_event_{event_id}")]
        ])

def get_description_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Пропустить")],
        ],
        resize_keyboard=True,
        one_time_keyboard=True
    )

def get_skip_edit_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Оставить без изменений")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def get_resume_registration_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="▶️ Продолжить регистрацию")]
        ],
        resize_keyboard=True,
        one_time_keyboard=True
    )


def get_gender_keyboard(edit_mode=False):
    keyboard = [
        [KeyboardButton(text="Муж"), KeyboardButton(text="Жен")],
        [KeyboardButton(text="Пропустить")]
    ]
    if edit_mode:
        keyboard.append([KeyboardButton(text="Оставить без изменений")])
    
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=True
    )


def get_region_keyboard(edit_mode=False):
    """Динамическая клавиатура регионов из БД"""
    regions = get_all_regions()
    kb = [[KeyboardButton(text=region)] for region in regions]
    
    if edit_mode:
        kb.append([KeyboardButton(text="Оставить без изменений")])
    
    return ReplyKeyboardMarkup(
        keyboard=kb,
        resize_keyboard=True,
        one_time_keyboard=True
    )

def get_interests_keyboard(selected: list[str] = [], edit_mode=False) -> InlineKeyboardMarkup:
    """Динамическая inline-клавиатура интересов из БД"""
    interests = get_all_interests()
    inline_kb = InlineKeyboardMarkup(inline_keyboard=[])
    for interest in interests:
        text = f"✅ {interest}" if interest in selected else interest
        inline_kb.inline_keyboard.append([
            InlineKeyboardButton(text=text, callback_data=interest)
        ])
    
    buttons_row = [InlineKeyboardButton(text="Готово", callback_data="done")]
    if edit_mode:
        buttons_row.append(InlineKeyboardButton(text="Оставить без изменений", callback_data="keep_current"))
    
    inline_kb.inline_keyboard.append(buttons_row)
    return inline_kb

def get_photo_keyboard(edit_mode=False):
    keyboard = [[KeyboardButton(text="Пропустить")]]
    
    if edit_mode:
        keyboard.append([KeyboardButton(text="Оставить без изменений")])
    
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=True
    )


def get_admin_menu_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📥 Загрузить списки")],
            [KeyboardButton(text="📊 Отчет по пользователям")],
            [KeyboardButton(text="📅 Отчет по мероприятиям")],
        ],
        resize_keyboard=True
    )


def get_user_main_menu():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👤 Мой профиль")],
            [KeyboardButton(text="💬 Общение")],
            [KeyboardButton(text="🎉 Мероприятия")],
            [KeyboardButton(text="❓ Помощь")],
        ],
        resize_keyboard=True,
        one_time_keyboard=False  
    )

def get_events_menu_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Список мероприятий")],
            [KeyboardButton(text="Мои мероприятия")],
            [KeyboardButton(text="Создать мероприятие")],
            [KeyboardButton(text="Назад")],
        ],
        resize_keyboard=True
    )


async def get_event_card_text(event: dict):
    """Формирует текст карточки мероприятия"""
    text = f"<b>{event['name']}</b>\n"
    text += f"📅 {event['date']} в {event['time']}\n"
    if event['address']:
        text += f"📍 {event['address']}\n"
    if event['interests']:
        text += f"❤️ {event['interests']}\n"
    if event['description']:
        text += f"\n{event['description']}\n"
    text += f"\nОрганизатор: +{event['organizer_phone']}\n"
    return text

async def find_potential_friends(organizer_phone: str, interests: list[str] = None):
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        query = """
            SELECT number, tg_id, name, surname, age, gender, region, interests
            FROM users
            WHERE number != ?
            AND registered = 1
            AND tg_id IS NOT NULL  -- только те, кто заходил в бот
        """
        params = [organizer_phone]

        if interests:
            query += " AND ("
            conditions = []
            for interest in interests:
                conditions.append("interests LIKE ?")
                params.append(f"%{interest}%")
            query += " OR ".join(conditions) + ")"

        c.execute(query, params)
        rows = c.fetchall()

    friends = []
    for row in rows:
        friends.append({
            "phone": row[0],
            "tg_id": row[1],         
            "name": row[2] or "—",
            "surname": row[3] or "",
            "age": row[4],
            "gender": row[5],
            "region": row[6],
            "interests": row[7].split(",") if row[7] else []
        })

    if interests:
        friends.sort(
            key=lambda f: len(set(interests) & set(f["interests"])),
            reverse=True
        )

    return friends[:20]

def is_valid_name(text: str) -> bool:
    return bool(re.match(r'^[a-zA-Zа-яА-ЯёЁ]+$', text))

def is_valid_age(text: str) -> bool:
    return text.isdigit() and 0 < int(text) < 120  


class UserMiddleware(BaseMiddleware):
    async def __call__(self, handler, event: TelegramObject, data: dict):
        if isinstance(event, Message):
            user_id = event.from_user.id
        elif isinstance(event, types.CallbackQuery):
            user_id = event.from_user.id
        else:
            return await handler(event, data)

        try:
            with sqlite3.connect(DB_PATH) as conn:
                c = conn.cursor()
                c.execute("""
                    SELECT number, role, registered, name, surname, gender, age,
                        region, interests, photo_file_id, document_file_id, 
                        location_lat, location_lon
                    FROM users WHERE tg_id = ?
                """, (user_id,))
                row = c.fetchone()

            if row:
                columns = [
                    "number", "role", "registered", "name", "surname", "gender",
                    "age", "region", "interests", "photo_file_id", "document_file_id",
                    "location_lat", "location_lon"
                ]
                user_data = dict(zip(columns, row))
                user_data["registered"] = bool(user_data["registered"])
                data["user"] = user_data
            else:
                data["user"] = None
        except Exception as e:
            logging.error(f"Middleware error for user {user_id}: {e}")
            data["user"] = None

        return await handler(event, data)
    

storage = MemoryStorage()
dp = Dispatcher(storage=storage)

dp.message.middleware(UserMiddleware())
dp.callback_query.middleware(UserMiddleware())


def get_start_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="Запустить")]],
        resize_keyboard=True,
        one_time_keyboard=True,
    )

def get_contact_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Предоставить номер", request_contact=True)]
        ],
        resize_keyboard=True,
        one_time_keyboard=True,
    )



@dp.message(CreateEvent.name)
async def process_event_name(message: Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer("Название не может быть пустым.")
        return
    await state.update_data(name=name)
    await state.set_state(CreateEvent.date)
    await message.answer("Введите дату начала (ДД.ММ.ГГГГ):")


@dp.message(CreateEvent.date)
async def process_event_date(message: Message, state: FSMContext):
    date_str = message.text.strip()

    if not re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_str):
        await message.answer("Неверный формат. Пример: 15.03.2025")
        return

    try:
        event_date = datetime.strptime(date_str, "%d.%m.%Y")
    except ValueError:
        await message.answer("Такая дата не существует. Попробуйте ещё раз.")
        return

    today = datetime.now().date()
    if event_date.date() < today:
        await message.answer("Нельзя создавать мероприятие в прошлом 😅")
        return

    await state.update_data(date=date_str, event_date_obj=event_date)
    await state.set_state(CreateEvent.time)
    await message.answer("Введите время начала (ЧЧ:ММ):")

@dp.message(CreateEvent.time)
async def process_event_time(message: Message, state: FSMContext):
    time_str = message.text.strip()

    if not re.match(r'^\d{2}:\d{2}$', time_str):
        await message.answer("Неверный формат. Пример: 18:30")
        return

    try:
        hours, minutes = map(int, time_str.split(":"))
        if not (0 <= hours <= 23 and 0 <= minutes <= 59):
            raise ValueError
    except ValueError:
        await message.answer("Такого времени не бывает. Попробуйте ещё раз.")
        return

    data = await state.get_data()
    event_date = data.get("event_date_obj")

    now = datetime.now()
    if event_date.date() == now.date():
        event_datetime = datetime.combine(event_date.date(), datetime.strptime(time_str, "%H:%M").time())
        if event_datetime <= now:
            await message.answer("Нельзя создавать мероприятие в прошлом или в текущий момент.")
            return

    await state.update_data(time=time_str)
    await state.set_state(CreateEvent.interests)
    await message.answer(
        "Выберите интересы мероприятия (можно несколько):",
        reply_markup=get_interests_keyboard([], edit_mode=False)
    )


@dp.callback_query(CreateEvent.interests)
async def process_event_interests(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    interests = data.get('interests', [])

    if callback.data == "done":
        if not interests:
            await callback.message.answer("Укажите хотя бы один интерес.")
            return
        await state.update_data(interests=interests)
        await state.set_state(CreateEvent.address)
        await callback.message.answer("Введите адрес мероприятия:")
        await callback.answer()
        return

    if callback.data in interests:
        interests.remove(callback.data)
    else:
        interests.append(callback.data)

    await state.update_data(interests=interests)
    await callback.message.edit_reply_markup(reply_markup=get_interests_keyboard(interests))
    await callback.answer()


@dp.message(CreateEvent.address)
async def process_event_address(message: Message, state: FSMContext):
    address = message.text.strip()
    await state.update_data(address=address)
    await state.set_state(CreateEvent.description)
    await message.answer(
        "Введите описание мероприятия (можно пропустить):",
        reply_markup=get_description_keyboard()
    )

@dp.message(CreateEvent.description, F.text == "Пропустить")
async def skip_event_description(message: Message, state: FSMContext):
    await state.update_data(description=None)  
    await state.set_state(CreateEvent.photo)
    await message.answer(
        "Загрузите фото мероприятия (jpg, jpeg, png) или пропустите:",
        reply_markup=get_photo_keyboard()
    )

@dp.message(CreateEvent.description)
async def process_event_description(message: Message, state: FSMContext):
    description = message.text.strip()
    await state.update_data(description=description)
    await state.set_state(CreateEvent.photo)
    await message.answer(
        "Загрузите фото мероприятия (jpg, jpeg, png) или пропустите:",
        reply_markup=get_photo_keyboard()
    )


@dp.message(CreateEvent.photo, F.photo)
async def process_event_photo_media(message: Message, state: FSMContext):
    photo = message.photo[-1]
    await state.update_data(photo_file_id=photo.file_id, document_file_id=None)
    await state.set_state(CreateEvent.invite_friends)
    await show_invite_friends_list(message, state)

@dp.message(CreateEvent.photo, F.document)
async def process_event_photo_document(message: Message, state: FSMContext):
    doc = message.document
    if not doc.mime_type or not doc.mime_type.startswith("image/"):
        await message.answer("🚫 Файл не является изображением.")
        return
    if not doc.file_name.lower().endswith((".jpg", ".jpeg", ".png")):
        await message.answer("🚫 Поддерживаются только JPG, JPEG, PNG.")
        return
    await state.update_data(document_file_id=doc.file_id, photo_file_id=None)
    await state.set_state(CreateEvent.invite_friends)
    await show_invite_friends_list(message, state)


@dp.message(CreateEvent.photo, F.text == "Пропустить")
async def process_event_photo_skip(message: Message, state: FSMContext):
    await state.update_data(photo_file_id=None, document_file_id=None)
    await state.set_state(CreateEvent.invite_friends)
    await show_invite_friends_list(message, state)


@dp.message(CreateEvent.photo)
async def process_event_photo_invalid(message: Message, state: FSMContext):
    await message.answer(
        "🚫 Отправьте фото (как изображение или файл JPG/PNG) "
        "или нажмите «Пропустить»"
    )



async def show_invite_friends_list(message: Message, state: FSMContext):
    data = await state.get_data()
    interests = data.get("interests", [])

    friends = await find_potential_friends(
        message.from_user.id,  
        interests
    )

    if not friends:
        await message.answer(
            "Пока нет подходящих друзей для приглашения 😔\n"
            "Можно продолжить без приглашения.",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="Продолжить без приглашения")]],
                resize_keyboard=True,
                one_time_keyboard=True
            )
        )
        return

    text = "Выберите друзей для приглашения:\n\n"
    kb = InlineKeyboardMarkup(inline_keyboard=[])

    for friend in friends:
        name = f"{friend['name']} {friend['surname']}".strip()
        age = friend['age'] if friend['age'] else "—"
        row = f"[{name}][{age}]"
        kb.inline_keyboard.append([
            InlineKeyboardButton(
                text=row,
                callback_data=f"invite_friend_{friend['tg_id']}"  
            )
        ])

    kb.inline_keyboard.append([
        InlineKeyboardButton(text="Пригласить всех", callback_data="invite_all"),
        InlineKeyboardButton(text="Продолжить без приглашения", callback_data="skip_invite")
    ])

    await message.answer(text, reply_markup=kb)


@dp.callback_query(lambda c: c.data.startswith("invite_friend_"))
async def invite_single_friend(callback: types.CallbackQuery, state: FSMContext, user: dict | None):
    if user is None:
        await callback.answer("Сессия истекла", show_alert=True)
        return

    friend_tg_id = int(callback.data.split("_")[2])

    data = await state.get_data()
    event_name = data.get("name", "Мероприятие")

    try:
        await callback.bot.send_message(
            friend_tg_id,
            f"Привет! {user['name']} пригласил тебя на мероприятие «{event_name}»!\n"
            f"Дата: {data.get('date')}, время: {data.get('time')}\n"
            f"Адрес: {data.get('address', 'не указан')}\n"
            "Присоединяйся! 🎉\n\n"
            "Чтобы посмотреть подробности — зайди в бот и нажми «Мероприятия» → «Список мероприятий»"
        )
        await callback.answer("Приглашение отправлено!", show_alert=True)

        await callback.bot.send_message(
            callback.from_user.id,
            f"Приглашение успешно отправлено пользователю с tg_id {friend_tg_id}!"
        )
    except Exception as e:
        logging.error(f"Ошибка отправки приглашения tg_id={friend_tg_id}: {e}")
        await callback.answer("Не удалось отправить приглашение (пользователь заблокировал бота?)", show_alert=True)


@dp.callback_query(F.data == "invite_all")
async def invite_all_friends(callback: types.CallbackQuery, state: FSMContext, user: dict | None):
    if user is None:
        await callback.answer("Сессия истекла", show_alert=True)
        return

    data = await state.get_data()
    interests = data.get("interests", [])
    event_name = data.get("name", "Мероприятие")

    friends = await find_potential_friends(user["number"], interests)

    sent_count = 0
    failed_count = 0
    for friend in friends:
        try:
            await callback.bot.send_message(
                friend["tg_id"],
                f"Привет! {user['name']} пригласил тебя на мероприятие «{event_name}»!\n"
                f"Дата: {data.get('date')}, время: {data.get('time')}\n"
                "Присоединяйся! 🎉"
            )
            sent_count += 1
        except Exception as e:
            logging.warning(f"Не удалось пригласить {friend['phone']}: {e}")
            failed_count += 1

    await callback.answer(f"Приглашено {sent_count} из {len(friends)} друзей!", show_alert=True)

    await callback.bot.send_message(
        callback.from_user.id,
        f"Приглашено {sent_count} из {len(friends)} друзей (не удалось: {failed_count})"
    )

    await state.set_state(CreateEvent.confirm)
    await show_event_preview(callback.message, state)
    try:
        await callback.message.delete()
    except:
        pass


@dp.callback_query(F.data == "skip_invite")
async def skip_invite(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(CreateEvent.confirm)
    await show_event_preview(callback.message, state)
    try:
        await callback.message.delete()
    except:
        pass
    await callback.answer("Приглашение пропущено")

async def show_event_preview(message: Message, state: FSMContext):
    data = await state.get_data()
    text = f"<b>{data['name']}</b>\n"
    text += f"Дата: {data['date']}\n"
    text += f"Время: {data['time']}\n"
    if data.get("interests"):
        text += f"Интересы: {', '.join(data['interests'])}\n"
    if data.get("address"):
        text += f"Адрес: {data['address']}\n"
    if data.get("description"):
        text += f"\n{data['description']}\n"
        

    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Сохранить")],
            [KeyboardButton(text="Отмена")],
        ],
        resize_keyboard=True
    )

    if data.get("photo_file_id"):
        await message.answer_photo(
            photo=data["photo_file_id"],
            caption=text,
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
    elif data.get("document_file_id"):
        await message.answer_document(
            document=data["document_file_id"],
            caption=text,
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
    else:
        await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)


@dp.message(CreateEvent.confirm)
async def process_event_confirm(message: Message, state: FSMContext, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сессия истекла. Начните заново.")
        await state.clear()
        return

    text = message.text.strip()

    if text == "Сохранить":
        data = await state.get_data()

        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            try:
                c.execute("""
                    INSERT INTO events (
                        organizer_phone, name, date, time, interests, address,
                        description, photo_file_id, document_file_id
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    user["number"],
                    data.get("name"),
                    data.get("date"),
                    data.get("time"),
                    ','.join(data.get("interests", [])) if data.get("interests") else None,
                    data.get("address"),
                    data.get("description"),
                    data.get("photo_file_id"),
                    data.get("document_file_id")
                ))
                conn.commit()
                await message.answer("Мероприятие успешно создано! 🎉", reply_markup=get_user_main_menu())
            except Exception as e:
                logging.error(f"Ошибка сохранения мероприятия: {e}")
                await message.answer("Ошибка при сохранении. Попробуйте позже.")
            finally:
                await state.clear()

    elif text == "Отмена":
        await state.clear()
        await message.answer("Создание отменено.", reply_markup=get_user_main_menu())

    else:
        await message.answer("Пожалуйста, выберите «Сохранить» или «Отмена».")


@dp.message(CreateEvent.confirm, F.text == "Сохранить")
async def save_event(message: Message, state: FSMContext, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сначала зарегистрируйтесь.")
        await state.clear()
        return

    data = await state.get_data()

    try:
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            c.execute("""
                INSERT INTO events (
                    organizer_phone, name, date, time, interests, address, 
                    description, photo_file_id, document_file_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                user["number"], data.get("name"), data.get("date"), data.get("time"),
                ','.join(data.get("interests", [])) if data.get("interests") else None,
                data.get("address"), data.get("description"),
                data.get("photo_file_id"), data.get("document_file_id")
            ))
            conn.commit() 
        
        await message.answer("Мероприятие успешно создано! 🎉", 
                           reply_markup=get_user_main_menu())
    except sqlite3.Error as e:
        logging.error(f"Database error при создании мероприятия: {e}")
        await message.answer("😔 Произошла ошибка при создании мероприятия. Попробуйте позже.")
    except Exception as e:
        logging.error(f"Unexpected error при создании мероприятия: {e}")
        await message.answer("😔 Непредвиденная ошибка. Обратитесь к администратору.")
    finally:
        await state.clear()


@dp.message(CreateEvent.confirm, F.text == "Отмена")
async def cancel_event_creation(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Создание мероприятия отменено.", reply_markup=get_user_main_menu())


@dp.message(F.text == "Создать мероприятие")
async def start_create_event(message: Message, state: FSMContext, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сначала зарегистрируйтесь.")
        return

    await state.set_state(CreateEvent.name)
    await message.answer(
        "Введите название мероприятия:",
        reply_markup=types.ReplyKeyboardRemove()
    )


@dp.message(F.text == "▶️ Продолжить регистрацию")
async def resume_registration(message: Message, state: FSMContext, user: dict | None):
    if not user:
        await message.answer("Ошибка. Пользователь не найден.")
        return

    if user["registered"]:
        await message.answer(
            "Регистрация уже завершена ✅",
            reply_markup=get_user_main_menu()
        )
        return

    await state.update_data(
        phone=user["number"]
    )

    await state.set_state(Registration.name)

    await message.answer(
        "Продолжаем регистрацию.\nВведите ваше имя:",
        reply_markup=types.ReplyKeyboardRemove()
    )


@dp.message(CommandStart())
async def cmd_start(message: Message, user: dict | None):
    if user:
        if user["role"] == "admin":
            await message.answer(
                "Добро пожаловать в админ-панель! 👑",
                reply_markup=get_admin_menu_keyboard()
            )
            return

        if user["registered"]:
            await message.answer(
                f"С возвращением, {user['name'] or 'пользователь'}! 👋",
                reply_markup=get_user_main_menu()
            )
            return

        await message.answer("Ваша регистрация не завершена. Продолжим?", reply_markup=get_resume_registration_keyboard())
    else:
        welcome = (
            "Что может этот бот:\n\n"
            "• Организовывать мероприятия\n"
            "• Искать участников\n"
            "• Общаться по интересам\n\n"
            "Нажмите «Запустить»"
        )
        await message.answer(welcome, reply_markup=get_start_keyboard())

@dp.message(lambda m: m.text == "Запустить")
async def btn_launch(message: Message, user: dict | None):
    if user:
        if user["role"] == "admin":
            await message.answer(
                "Добро пожаловать в админ-панель! 👑",
                reply_markup=get_admin_menu_keyboard()
            )
            return

        if user["registered"]:
            await message.answer(
                f"С возвращением, {user['name'] or 'пользователь'}! 👋",
                reply_markup=get_user_main_menu()
            )
            return

        await message.answer("Ваша регистрация не завершена. Продолжим?", reply_markup=get_resume_registration_keyboard())
    else:
        text = "Предоставьте номер телефона для регистрации / авторизации"
        await message.answer(text, reply_markup=get_contact_keyboard())

@dp.message(F.contact)
async def process_contact(message: Message, state: FSMContext, user: dict | None):
    phone = message.contact.phone_number.strip()
    tg_id = message.from_user.id

    if user is not None:
        if user["role"] == "admin":
            await message.answer(
                "Добро пожаловать в админ-панель!",
                reply_markup=get_admin_menu_keyboard()
            )
            return

        if user["registered"]:
            await message.answer(
                f"С возвращением, {user['name'] or 'пользователь'}!",
                reply_markup=get_user_main_menu()
            )
            return
        else:
            await message.answer("Давайте завершим регистрацию.")
    else:
        success = register_phone(phone, tg_id)
        if success:
            await message.answer(f"Номер {phone} добавлен. Заполняем профиль.")
        else:
            await message.answer("Ошибка при сохранении номера.")
            return

    await state.update_data(phone=phone)
    await state.set_state(Registration.name)
    await message.answer("Введите ваше Имя:", reply_markup=types.ReplyKeyboardRemove())

@dp.message(F.text.in_({"отмена", "cancel", "Отмена", "/cancel"}))
async def cmd_cancel(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is None:
        return
    await state.clear()
    await message.answer("Действие отменено.", reply_markup=get_user_main_menu())



@dp.message(F.text == "🎉 Мероприятия")
async def events_menu(message: Message, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сначала завершите регистрацию.")
        return

    await message.answer(
        "Раздел мероприятий 🎉",
        reply_markup=get_events_menu_keyboard()
    )

@dp.message(F.text == "Список мероприятий")
async def show_all_events(message: Message, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сначала зарегистрируйтесь.")
        return

    phone = user["number"]

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            SELECT 
                e.id, e.name, e.date, e.time, e.address, e.interests, 
                e.description, e.organizer_phone,
                CASE WHEN ep.participant_phone IS NOT NULL THEN 1 ELSE 0 END as is_participant
            FROM events e
            LEFT JOIN event_participants ep 
                ON e.id = ep.event_id AND ep.participant_phone = ?
            ORDER BY e.created_at DESC
        """, (phone,))
        events = c.fetchall()

    if not events:
        await message.answer("Пока нет ни одного мероприятия 😔")
        return

    for event in events:
        event_dict = {
            "id": event[0],
            "name": event[1],
            "date": event[2],
            "time": event[3],
            "address": event[4],
            "interests": event[5],
            "description": event[6],
            "organizer_phone": event[7],
            "is_participant": bool(event[8])
        }
        text = await get_event_card_text(event_dict)
        
        kb = get_event_card_keyboard_optimized(
            event_dict["id"], 
            phone, 
            event_dict["organizer_phone"],
            event_dict["is_participant"]
        )

        await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)


@dp.message(F.text == "Мои мероприятия")
async def show_my_events(message: Message, user: dict | None):
    if user is None or not user["registered"]:
        await message.answer("Сначала зарегистрируйтесь.")
        return

    phone = user["number"]
    
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()

        c.execute("""
            SELECT 
                e.id, e.name, e.date, e.time, e.address, e.interests, 
                e.description, e.organizer_phone,
                1 as is_organizer,
                0 as is_participant
            FROM events e
            WHERE e.organizer_phone = ?
            ORDER BY e.created_at DESC
        """, (phone,))
        my_organized = c.fetchall()

        c.execute("""
            SELECT 
                e.id, e.name, e.date, e.time, e.address, e.interests, 
                e.description, e.organizer_phone,
                0 as is_organizer,
                1 as is_participant
            FROM events e
            JOIN event_participants ep ON e.id = ep.event_id
            WHERE ep.participant_phone = ?
            ORDER BY e.created_at DESC
        """, (phone,))
        my_participated = c.fetchall()

    if not my_organized and not my_participated:
        await message.answer("У вас пока нет мероприятий 😔")
        return

    if my_organized:
        await message.answer("📌 Мои мероприятия (организатор):")
        for event in my_organized:
            event_dict = {
                "id": event[0], "name": event[1], "date": event[2], 
                "time": event[3], "address": event[4], "interests": event[5], 
                "description": event[6], "organizer_phone": event[7]
            }
            text = await get_event_card_text(event_dict)
            await message.answer(text, parse_mode=ParseMode.HTML)
            
    if my_participated:
        await message.answer("🎟️ Мои мероприятия (участник):")
        for event in my_participated:
            event_dict = {
                "id": event[0], "name": event[1], "date": event[2], 
                "time": event[3], "address": event[4], "interests": event[5], 
                "description": event[6], "organizer_phone": event[7],
                "is_participant": bool(event[9])
            }
            text = await get_event_card_text(event_dict)
            kb = get_event_card_keyboard_optimized(
                event_dict["id"], phone, event_dict["organizer_phone"], 
                event_dict["is_participant"]
            )
            await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)


@dp.callback_query(lambda c: c.data.startswith("join_event_"))
async def join_event(callback: types.CallbackQuery, user: dict | None):
    if user is None or not user["registered"]:
        await callback.answer("Сначала зарегистрируйтесь.", show_alert=True)
        return

    event_id = int(callback.data.split("_")[2])
    phone = user["number"]

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        try:
            c.execute("""
                INSERT INTO event_participants (event_id, participant_phone)
                VALUES (?, ?)
            """, (event_id, phone))
            conn.commit()
            await callback.answer("Вы успешно записались! 🎉", show_alert=True)
            await callback.message.edit_reply_markup(reply_markup=None)
        except sqlite3.IntegrityError:
            await callback.answer("Вы уже участвуете в этом мероприятии!", show_alert=True)


@dp.callback_query(lambda c: c.data.startswith("leave_event_"))
async def leave_event(callback: types.CallbackQuery, user: dict | None):
    if user is None or not user["registered"]:
        await callback.answer("Сначала зарегистрируйтесь.", show_alert=True)
        return

    event_id = int(callback.data.split("_")[2])
    phone = user["number"]

    try:
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            
            c.execute("""
                SELECT e.organizer_phone 
                FROM events e
                WHERE e.id = ?
            """, (event_id,))
            result = c.fetchone()
            
            if not result:
                await callback.answer("Мероприятие не найдено!", show_alert=True)
                return
            
            organizer_phone = result[0]
            
            c.execute("""
                DELETE FROM event_participants 
                WHERE event_id = ? AND participant_phone = ?
            """, (event_id, phone))
            
            if c.rowcount == 0:
                await callback.answer("Вы не участвуете в этом мероприятии!", show_alert=True)
                return
            
            conn.commit()

        await callback.answer("Вы отказались от участия", show_alert=True)
        
        new_kb = get_event_card_keyboard_optimized(event_id, phone, organizer_phone, False)
        if new_kb:
            await callback.message.edit_reply_markup(reply_markup=new_kb)
        else:
            await callback.message.edit_reply_markup(reply_markup=None)
            
    except sqlite3.Error as e:
        logging.error(f"Database error в leave_event: {e}")
        await callback.answer("😔 Ошибка при отказе от участия", show_alert=True)
    except Exception as e:
        logging.error(f"Unexpected error в leave_event: {e}")
        await callback.answer("😔 Непредвиденная ошибка", show_alert=True)


@dp.message(F.text == "Назад")
async def back_to_main(message: Message):
    await message.answer("Главное меню", reply_markup=get_user_main_menu())


@dp.message(F.text == "📥 Загрузить списки")
async def admin_load_lists(message: Message, state: FSMContext, user: dict | None):
    if user is None or user["role"] != "admin":
        await message.answer("Эта функция доступна только администраторам.")
        return

    await state.set_state(AdminLoad.waiting_excel)
    await message.answer(
        "📎 Загрузите Excel-файл:\n\n"
        "• Столбец A — Интересы\n"
        "• Столбец B — Регионы"
    )

@dp.message(AdminLoad.waiting_excel, F.document)
async def admin_process_excel(message: Message, state: FSMContext, user: dict | None):
    doc = message.document

    if user is None or user["role"] != "admin":
        await message.answer("Доступ запрещён.")
        return

    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await message.answer("🚫 Поддерживаются только Excel-файлы")
        return

    file_id = uuid.uuid4()
    file_ext = os.path.splitext(doc.file_name)[1]  
    file_path = f"/tmp/{file_id}{file_ext}"

    try:
        file = await message.bot.get_file(doc.file_id)
        await message.bot.download_file(file.file_path, file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        interests, regions = [], []
        for row in ws.iter_rows(values_only=True):
            if row and row[0]:
                interests.append(str(row[0]).strip())
            if row and len(row) > 1 and row[1]:
                regions.append(str(row[1]).strip())

        if not interests and not regions:
            await message.answer("🚫 Файл пуст или не содержит данных")
            return

        replace_interests(interests)
        replace_regions(regions)

        await state.clear()
        await message.answer("✅ Списки успешно обновлены", 
                           reply_markup=get_admin_menu_keyboard())
    except Exception as e:
        logging.error(f"Excel processing error: {e}")
        await message.answer("🚫 Ошибка обработки файла")
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

@dp.message(Registration.name)
async def reg_name(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    name = message.text.strip()
    
    if edit_mode and name == "Оставить без изменений":
        name = data.get("name")
    else:
        if not is_valid_name(name):
            await message.answer("🚫 Не похоже на имя. Только буквы. Попробуйте еще раз.")
            return

    try:
        await state.update_data(name=name)

        if edit_mode:
            current = data.get("surname", "не указано")
            await message.answer(
                f"Текущая фамилия: {current}\nВведите новую фамилию или оставьте без изменений:",
                reply_markup=get_skip_edit_keyboard()
            )
        else:
            await message.answer("Введите вашу фамилию:")

        await state.set_state(Registration.surname)
    except Exception as e:
        logging.error(f"Error in reg_name for user {message.from_user.id}: {e}")
        await message.answer("😔 Произошла ошибка. Попробуйте ещё раз или обратитесь к администратору.")
        await state.clear()


@dp.message(Registration.surname)
async def reg_surname(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    surname = message.text.strip()
    
    if edit_mode and surname == "Оставить без изменений":
        surname = data.get("surname")
    else:
        if not is_valid_name(surname):
            await message.answer("🚫 Не похоже на фамилию. Только буквы. Попробуйте еще раз.")
            return

    await state.update_data(surname=surname)

    if edit_mode:
        current = data.get("gender") or "не указан"
        await message.answer(f"Текущий пол: {current}")

    await message.answer("Укажите пол:", reply_markup=get_gender_keyboard(edit_mode))  
    await state.set_state(Registration.gender)

@dp.message(Registration.gender)
async def reg_gender(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    gender = message.text.strip()
    
    if edit_mode and gender == "Оставить без изменений":
        gender = data.get("gender")
    elif gender == "Пропустить":
        gender = None
    elif gender not in ["Муж", "Жен"]:
        await message.answer("🚫 Выберите из кнопок или пропустите.")
        return

    await state.update_data(gender=gender)

    if edit_mode:
        current = data.get("age", "не указан")
        await message.answer(
            f"Текущий возраст: {current}\nВведите новый возраст или оставьте без изменений:",
            reply_markup=get_skip_edit_keyboard()
        )
    else:
        await message.answer("Укажите возраст:", reply_markup=types.ReplyKeyboardRemove())

    await state.set_state(Registration.age)

@dp.message(Registration.age)
async def reg_age(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    age = message.text.strip()
    
    if edit_mode and age == "Оставить без изменений":
        age = data.get("age")  
    else:
        if not is_valid_age(age):
            await message.answer("🚫 Не похоже на возраст. Только цифры. Попробуйте еще раз.")
            return
        age = int(age)

    await state.update_data(age=age)

    if edit_mode:
        current = data.get("region", "не указан")
        await message.answer(f"Текущий регион: {current}")

    await message.answer("Выберите ваш регион:", reply_markup=get_region_keyboard(edit_mode))  
    await state.set_state(Registration.region)

@dp.message(Registration.region)
async def reg_region(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    region = message.text.strip()
    
    if edit_mode and region == "Оставить без изменений":
        region = data.get("region")  
    else:
        regions = get_all_regions()
        if region not in regions:
            await message.answer("🚫 Выберите из списка.")
            return

    await state.update_data(region=region)

    if edit_mode:
        current = ", ".join(data.get("interests", [])) or "не указаны"
        await message.answer(f"Текущие интересы: {current}")

    if not edit_mode:
        await state.update_data(interests=[])
    
    await message.answer(
        "Укажите ваши интересы (можно выбрать несколько):",
        reply_markup=get_interests_keyboard(data.get("interests", []), edit_mode)
    )
    await state.set_state(Registration.interests)

@dp.callback_query(Registration.interests)
async def reg_interests_callback(callback: types.CallbackQuery, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)
    interests = data.get('interests', [])

    if callback.data == "keep_current":
        await state.set_state(Registration.photo)
        
        if edit_mode:
            current = "есть" if data.get("photo_file_id") or data.get("document_file_id") else "нет"
            await callback.message.answer(f"Текущее фото: {current}")
        
        await callback.message.answer(
            "Загрузите фото (jpg, jpeg, png) или пропустите:",
            reply_markup=get_photo_keyboard(edit_mode) 
        )
        await callback.answer()
        return

    if callback.data == "done":
        if not interests:
            await callback.answer("🚫 Укажите хотя бы один интерес.")
            return

        await state.update_data(interests=interests)

        if edit_mode:
            current = "есть" if data.get("photo_file_id") or data.get("document_file_id") else "нет"
            await callback.message.answer(f"Текущее фото: {current}")

        await state.set_state(Registration.photo)
        await callback.message.answer(
            "Загрузите фото (jpg, jpeg, png) или пропустите:",
            reply_markup=get_photo_keyboard(edit_mode)  
        )
        await callback.answer()
        return

    if callback.data in interests:
        interests.remove(callback.data)
    else:
        interests.append(callback.data)

    await state.update_data(interests=interests)
    await callback.message.edit_reply_markup(reply_markup=get_interests_keyboard(interests, edit_mode))
    await callback.answer()


@dp.message(Registration.photo, F.text == "Оставить без изменений")
async def reg_photo_keep(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)
    
    if edit_mode:
        current = "есть" if data.get("location_lat") else "нет"
        await message.answer(f"Текущее местоположение: {current}")

    await state.set_state(Registration.location)
    await ask_user_location(message, edit_mode)


@dp.message(Registration.photo, F.photo)
async def reg_photo_media(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    photo = message.photo[-1]
    await state.update_data(photo_file_id=photo.file_id, document_file_id=None)

    if edit_mode:
        current = "есть" if data.get("location_lat") else "нет"
        await message.answer(f"Текущее местоположение: {current}")

    await state.set_state(Registration.location)
    await ask_user_location(message, edit_mode)


@dp.message(Registration.photo, F.document)
async def reg_photo_document(message: Message, state: FSMContext, user: dict | None):
    doc = message.document

    if not doc.mime_type or not doc.mime_type.startswith("image/"):
        await message.answer("🚫 Файл не является изображением.")
        return

    if not doc.file_name.lower().endswith((".jpg", ".jpeg", ".png")):
        await message.answer("🚫 Поддерживаются только JPG, JPEG, PNG.")
        return

    await state.update_data(document_file_id=doc.file_id, photo_file_id=None)

    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    if edit_mode:
        current = "есть" if data.get("location_lat") else "нет"
        await message.answer(f"Текущее местоположение: {current}")

    await state.set_state(Registration.location)
    await ask_user_location(message, edit_mode)


@dp.message(Registration.photo, F.text == "Пропустить")
async def reg_photo_skip(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    await state.update_data(photo_file_id=None, document_file_id=None)

    if edit_mode:
        current = "есть" if data.get("location_lat") else "нет"
        await message.answer(f"Текущее местоположение: {current}")

    await state.set_state(Registration.location)
    await ask_user_location(message, edit_mode)


@dp.message(Registration.photo)
async def reg_photo_invalid(message: Message, state: FSMContext, user: dict | None):
    await message.answer(
        "🚫 Отправьте фото (как изображение или файл JPG/PNG) "
        "или нажмите «Пропустить»"
    )

def get_location_keyboard(edit_mode=False):
    keyboard = [
        [KeyboardButton(text="📱 Поделиться геолокацией", request_location=True)],
        [KeyboardButton(text="💻 Ручной ввод координат")],
        [KeyboardButton(text="Пропустить")]
    ]
    
    if edit_mode:
        keyboard.append([KeyboardButton(text="Оставить без изменений")])
    
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=True
    )


async def ask_user_location(message: Message, edit_mode=False):
    await message.answer(
        "Укажите ваше местоположение:\n\n"
        "📱 На телефоне — нажмите «Поделиться геолокацией». "
        "Для этого на устройстве должна быть включена геолокация.\n"
        "💻 На ПК — выберите «Ручной ввод» и введите координаты вручную "
        "(широта, долгота, например: 55.7558, 37.6173).\n"
        "Можно также нажать «Пропустить», если не хотите указывать местоположение.",
        reply_markup=get_location_keyboard(edit_mode) 
    )


@dp.message(Registration.location, F.text == "Оставить без изменений")
async def reg_location_keep(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)
    
    update_user_profile(data["phone"], data)
    await state.clear()

    text = "Профиль обновлён!" if edit_mode else "Регистрация завершена! Добро пожаловать 🎉"
    await message.answer(text, reply_markup=get_user_main_menu())

@dp.message(Registration.location, F.location)
async def reg_location_ok(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    await state.update_data(
        location_lat=message.location.latitude,
        location_lon=message.location.longitude
    )
    data = await state.get_data()
    update_user_profile(data["phone"], data)
    await state.clear()

    text = "Профиль обновлён!" if edit_mode else "Регистрация завершена! Добро пожаловать 🎉"
    await message.answer(text, reply_markup=get_user_main_menu())


@dp.message(Registration.location, F.text == "💻 Ручной ввод координат")
async def reg_location_manual_start(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    current = "есть" if edit_mode and data.get("location_lat") else "не указано"
    await message.answer(
        f"Текущее местоположение: {current}\n"
        "Введите координаты вручную в формате: широта, долгота\n"
        "Пример: 55.7558, 37.6173"
    )


@dp.message(Registration.location)
async def reg_location_manual_process(
    message: Message,
    state: FSMContext,
    user: dict | None
):
    if not message.text:
        await message.answer(
            "Отправь координаты в формате:\n"
            "55.7558, 37.6173\n"
            "или нажми «Пропустить»"
        )
        return

    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    text = message.text.strip()

    if text.lower() == "пропустить":
        await state.update_data(location_lat=None, location_lon=None)
        update_user_profile(data["phone"], data)
        await state.clear()

        text_msg = (
            "Профиль обновлён!"
            if edit_mode
            else "Регистрация завершена! Добро пожаловать 🎉"
        )
        await message.answer(text_msg, reply_markup=get_user_main_menu())
        return

    match = re.match(
        r'^\s*(-?\d+(\.\d+)?)\s*,\s*(-?\d+(\.\d+)?)\s*$',
        text
    )

    if not match:
        await message.answer(
            "🚫 Неверный формат.\n"
            "Пример: 55.7558, 37.6173\n"
            "Или нажмите «Пропустить»"
        )
        return

    lat = float(match.group(1))
    lon = float(match.group(3))

    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        await message.answer(
            "🚫 Координаты вне допустимого диапазона.\n"
            "Широта: от -90 до 90\n"
            "Долгота: от -180 до 180"
        )
        return

    await state.update_data(location_lat=lat, location_lon=lon)
    update_user_profile(data["phone"], data)
    await state.clear()

    text_msg = (
        "Профиль обновлён!"
        if edit_mode
        else "Регистрация завершена! Добро пожаловать 🎉"
    )
    await message.answer(text_msg, reply_markup=get_user_main_menu())

@dp.message(Registration.location, F.text == "Пропустить")
async def reg_location_skip(message: Message, state: FSMContext, user: dict | None):
    data = await state.get_data()
    edit_mode = data.get("edit_mode", False)

    await state.update_data(location_lat=None, location_lon=None)
    update_user_profile(data["phone"], data)
    await state.clear()

    text = "Профиль обновлён!" if edit_mode else "Регистрация завершена! Добро пожаловать 🎉"
    await message.answer(text, reply_markup=get_user_main_menu())


@dp.message(Registration.location)
async def reg_location_invalid(message: Message, state: FSMContext, user: dict | None):
    await message.answer(
        "🚫 Пожалуйста, выберите одну из кнопок:\n"
        "📱 Поделиться геолокацией\n"
        "💻 Ручной ввод координат\n"
        "Пропустить"
    )


@dp.message(F.text == "👤 Мой профиль")
async def show_my_profile(message: Message, user: dict | None):
    if user is None:
        await message.answer("Сначала зарегистрируйтесь.")
        return

    text = f"👤 <b>{user['name'] or '—'} {user['surname'] or ''}</b>\n"
    if user['age']:
        text += f"🎂 Возраст: {user['age']}\n"
    if user['gender']:
        text += f"🚻 Пол: {user['gender']}\n"
    if user['region']:
        text += f"📍 Регион: {user['region']}\n"
    if user['interests']:
        text += f"❤️ Интересы: {user['interests'].replace(',', ', ')}\n"

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✏️ Редактировать данные", callback_data="edit_profile")]
    ])

    if user['photo_file_id']:
        await message.answer_photo(
            photo=user['photo_file_id'],
            caption=text,
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
    elif user['document_file_id']:
        await message.answer_document(
            document=user['document_file_id'],
            caption=text,
            reply_markup=kb,
            parse_mode=ParseMode.HTML
        )
    else:
        await message.answer(text, reply_markup=kb, parse_mode=ParseMode.HTML)


@dp.callback_query(F.data == "edit_profile")
async def start_edit_profile(callback: types.CallbackQuery, state: FSMContext, user: dict | None):
    if user is None:
        await callback.message.answer("Ошибка. Пользователь не найден.")
        await callback.answer()
        return

    await state.update_data(
        phone=user["number"],
        edit_mode=True,  
        name=user["name"],
        surname=user["surname"],
        gender=user["gender"],
        age=user["age"],
        region=user["region"],
        interests=user["interests"].split(",") if user["interests"] else [],        
        photo_file_id=user["photo_file_id"],
        document_file_id=user["document_file_id"],
        location_lat=user["location_lat"],
        location_lon=user["location_lon"]
    )

    await state.set_state(Registration.name)

    current_name = user["name"] or "не указано"
    await callback.message.answer(
        f"Редактируем профиль.\nТекущее имя: {current_name}\nВведите новое имя или оставьте без изменений:",
        reply_markup=get_skip_edit_keyboard()  
    )
    await callback.answer()


async def main():
    init_db()
    init_admin_tables()

    if not get_all_regions():
        print("ВНИМАНИЕ: таблица regions пуста! Загрузите Excel-файл через админ-панель.")
    if not get_all_interests():
        print("ВНИМАНИЕ: таблица interests пуста! Загрузите Excel-файл через админ-панель.")

    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )

    logging.basicConfig(
        level=logging.INFO,
        stream=sys.stdout,
        format="%(asctime)s | %(levelname)-8s | %(name)s → %(message)s",
    )

    print("Бот запускается...")
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())