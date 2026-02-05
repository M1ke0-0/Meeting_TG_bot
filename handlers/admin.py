import logging
import uuid
import os
from openpyxl import load_workbook
from aiogram import Router, F, types
from aiogram.types import Message
from aiogram.fsm.context import FSMContext

from states.states import AdminLoad
from keyboards.builders import get_admin_menu_keyboard

from database import get_session
from database.repositories import InterestRepository, RegionRepository
from utils.excel import export_users_report, export_events_report

router = Router()

@router.message(F.text == "📥 Загрузить списки")
async def admin_load_lists(message: Message, state: FSMContext, user: dict | None):
    if user is None or user["role"] != "admin":
        return
    
    await message.answer(
        "Отправьте Excel-файл (.xlsx) с двумя листами:\n"
        "1. Interests (список интересов)\n"
        "2. Regions (список регионов)"
    )
    await state.set_state(AdminLoad.waiting_excel)


@router.message(AdminLoad.waiting_excel, F.document)
async def process_excel(message: Message, state: FSMContext, user: dict | None):
    if user is None or user["role"] != "admin":
        return

    doc = message.document
    if not doc.file_name.endswith('.xlsx'):
        await message.answer("Пожалуйста, отправьте файл .xlsx")
        return

    file_id = doc.file_id
    file = await message.bot.get_file(file_id)
    file_path = file.file_path
    
    temp_filename = f"temp_{uuid.uuid4()}.xlsx"
    await message.bot.download_file(file_path, temp_filename)

    try:
        wb = load_workbook(temp_filename)
        
        interests = []
        regions = []
        
        interest_sheet_names = ["Interests", "Интересы", "interests", "интересы"]
        region_sheet_names = ["Regions", "Регионы", "regions", "регионы"]
        
        interest_ws = None
        for name in interest_sheet_names:
            if name in wb.sheetnames:
                interest_ws = wb[name]
                break
        
        if interest_ws is None and len(wb.sheetnames) == 1:
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 1 and row[0]:
                    interests.append(str(row[0]).strip())
                if row and len(row) >= 2 and row[1]:
                    regions.append(str(row[1]).strip())
        else:
            if interest_ws:
                for row in interest_ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        interests.append(str(row[0]).strip())
            
            for name in region_sheet_names:
                if name in wb.sheetnames:
                    region_ws = wb[name]
                    for row in region_ws.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            regions.append(str(row[0]).strip())
                    break
        
        async with get_session() as session:
            interest_repo = InterestRepository(session)
            region_repo = RegionRepository(session)
            
            if interests:
                await interest_repo.replace_all(interests)
            if regions:
                await region_repo.replace_all(regions)

        await message.answer(
            f"✅ Обновлено:\n"
            f"Интересов: {len(interests)}\n"
            f"Регионов: {len(regions)}",
            reply_markup=get_admin_menu_keyboard()
        )
        
        
        try:
            await message.delete()
        except Exception:
            pass 
        
        await state.clear()

    except Exception as e:
        logging.error(f"Error processing Excel: {e}")
        await message.answer(f"Ошибка при обработке файла: {e}")
    finally:
        if os.path.exists(temp_filename):
            os.remove(temp_filename)


@router.message(F.text == "📊 Отчет по пользователям")
async def report_users(message: Message, user: dict | None):
    if user is None or user["role"] != "admin":
        return

    filename = f"users_report_{uuid.uuid4()}.xlsx"
    
    try:
       
        await export_users_report(filename)
        
        input_file = types.FSInputFile(filename)
        await message.answer_document(input_file, caption="Отчет по пользователям")
        
    except Exception as e:
        logging.error(f"Error generating users report: {e}")
        await message.answer(f"Ошибка генерации отчета: {e}")
    finally:
        if os.path.exists(filename):
            os.remove(filename)


@router.message(F.text == "📅 Отчет по мероприятиям")
async def report_events(message: Message, user: dict | None):
    if user is None or user["role"] != "admin":
        return

    filename = f"events_report_{uuid.uuid4()}.xlsx"
    
    try:
        await export_events_report(filename)
        
        input_file = types.FSInputFile(filename)
        await message.answer_document(input_file, caption="Отчет по мероприятиям")
        
    except Exception as e:
        logging.error(f"Error generating events report: {e}")
        await message.answer(f"Ошибка генерации отчета: {e}")
    finally:
        if os.path.exists(filename):
            os.remove(filename)
